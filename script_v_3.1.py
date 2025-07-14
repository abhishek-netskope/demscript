#!/usr/bin/env python3
"""
Netskope User Report Generator with Flexible Date Ranges & Full Debug Info
"""

# ================= CONFIGURATION =================
API_URL = "https://eliterhythm.goskope.com/api/v2/adem/users/getentities"
API_TOKEN = "12343212ss98765sa54a0s0s0d252576"
SORT_ORDER = "desc"
DEBUG_MODE = True
# =================================================

import requests
import json
import pandas as pd
from datetime import datetime, timedelta, timezone
import os
from typing import Dict, List, Optional
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
import logging
from tqdm import tqdm
from colorama import init, Fore, Style
from urllib.parse import urlparse

init(autoreset=True)
log_level = logging.DEBUG if DEBUG_MODE else logging.INFO
logging.basicConfig(level=log_level,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.FileHandler('netskope_debug.log'), logging.StreamHandler()])
logger = logging.getLogger(__name__)

class NetskopeAPIClient:
    def __init__(self, api_url: str, api_token: str):
        self.api_url = api_url
        self.api_token = api_token
        self.headers = {
            'Authorization': f'Bearer {api_token}',
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        }
        self.session = requests.Session()
        self.session.headers.update(self.headers)
        self.MAX_LIMIT = 100

    def get_users(self, limit=100, offset=0, start_time=None, end_time=None, sort_order="desc"):
        limit = min(limit, self.MAX_LIMIT)
        params = {
            "limit": limit,
            "offset": offset,
            "sortby": "expScore",
            "sortorder": sort_order
        }
        body = {}
        if start_time and end_time:
            body["starttime"] = int(start_time.timestamp())
            body["endtime"] = int(end_time.timestamp())
        logger.debug(f"API Request to {self.api_url} with params={params} and body={json.dumps(body)}")
        try:
            response = self.session.post(self.api_url, params=params, json=body)
            if response.status_code == 429:
                retry_after = int(response.headers.get('Retry-After', 60))
                logger.warning(f"Rate limited. Waiting {retry_after} seconds...")
                print(f"\n{Fore.YELLOW}⚠️  Rate limited. Waiting {retry_after} seconds...{Style.RESET_ALL}")
                time.sleep(retry_after)
                response = self.session.post(self.api_url, params=params, json=body)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"API request failed: {e}")
            raise

    def get_all_users_chunked(self, start_time, end_time, sort_order="desc"):
        all_users = []
        unique_users = {}
        duplicate_count = 0
        total_api_calls = 0
        delta = timedelta(hours=48)
        chunk_start = start_time
        total_chunks = int((end_time - start_time) / delta) + 1
        pbar = tqdm(desc="Time windows", total=total_chunks, unit="window", colour='green')
        while chunk_start < end_time:
            chunk_end = min(chunk_start + delta, end_time)
            print(f"{Fore.MAGENTA}\n🔎 Fetching chunk: {chunk_start} to {chunk_end}{Style.RESET_ALL}")
            offset = 0
            batch_count = 0
            while True:
                batch_count += 1
                total_api_calls += 1
                logger.info(f"Fetching batch {batch_count} in chunk ({chunk_start} - {chunk_end}), offset={offset}")
                batch = self.get_users(
                    limit=self.MAX_LIMIT,
                    offset=offset,
                    start_time=chunk_start,
                    end_time=chunk_end,
                    sort_order=sort_order
                )
                users = batch.get('users', [])
                if not users:
                    break
                batch_new = 0
                for user in users:
                    exp_score = user.get('expScore', 0)
                    if exp_score is None or exp_score <= 0:
                        continue
                    user_email = user.get('user', '').strip().lower()
                    if user_email:
                        if user_email not in unique_users:
                            unique_users[user_email] = user
                            all_users.append(user)
                            batch_new += 1
                        else:
                            duplicate_count += 1
                print(
                    f"{Fore.BLUE}[Chunk: {chunk_start.strftime('%Y-%m-%d %H:%M')} Batch {batch_count}] "
                    f"Unique users so far: {len(unique_users):,} (+{batch_new}), Duplicates: {duplicate_count:,}{Style.RESET_ALL}"
                )
                if len(users) < self.MAX_LIMIT:
                    break
                offset += self.MAX_LIMIT
                time.sleep(0.2)  # small delay for rate limits
            chunk_start = chunk_end
            pbar.update(1)
        pbar.close()
        print(f"{Fore.GREEN}✓ Total unique users: {len(unique_users):,}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}📊 Total duplicates skipped: {duplicate_count:,}{Style.RESET_ALL}")
        print(f"{Fore.CYAN}📊 Total API calls made: {total_api_calls}{Style.RESET_ALL}")
        return list(unique_users.values())

class UserDataProcessor:
    @staticmethod
    def process_users(users: List[Dict]) -> pd.DataFrame:
        print(f"\n{Fore.CYAN}🔄 Processing user data (expScore > 0)...{Style.RESET_ALL}")
        processed_data = []
        for user in users:
            exp_score = user.get('expScore', 0)
            if exp_score is None or exp_score <= 0:
                continue
            devices = user.get('devices', [])
            device_names = ', '.join([d.get('deviceName', '') for d in devices])
            device_classifications = ', '.join([d.get('deviceClassification', '') for d in devices])
            device_count = len(devices)
            applications = user.get('applications', [])
            app_names = ', '.join(applications) if applications else 'None'
            user_groups = user.get('userGroups', [])
            group_names = ' | '.join(user_groups) if user_groups else 'None'
            processed_user = {
                'User Email': user.get('user', ''),
                'Experience Score': exp_score,
                'Location': user.get('location', ''),
                'Device Count': device_count,
                'Device Names': device_names,
                'Device Classifications': device_classifications,
                'Applications Count': user.get('applicationsCount', 0),
                'Applications': app_names,
                'User Groups': group_names,
                'NPA Hosts': ', '.join(user.get('npaHosts', [])) if user.get('npaHosts') else 'None'
            }
            processed_data.append(processed_user)
        df = pd.DataFrame(processed_data)
        print(f"{Fore.GREEN}✓ Processed {len(df):,} user records{Style.RESET_ALL}")
        return df

    @staticmethod
    def create_group_aggregation(df: pd.DataFrame) -> pd.DataFrame:
        print(f"\n{Fore.CYAN}📊 Creating group-wise aggregations...{Style.RESET_ALL}")
        expanded_data = []
        for _, row in df.iterrows():
            groups = row['User Groups'].split(' | ') if row['User Groups'] != 'None' else ['No Group']
            for group in groups:
                expanded_data.append({
                    'Group': group.strip(),
                    'User': row['User Email'],
                    'Experience Score': row['Experience Score']
                })
        expanded_df = pd.DataFrame(expanded_data)
        group_agg = expanded_df.groupby('Group').agg({
            'User': 'count',
            'Experience Score': ['mean', 'min', 'max']
        }).round(2)
        group_agg.columns = ['User Count', 'Avg Score', 'Min Score', 'Max Score']
        group_agg = group_agg.reset_index()
        # Shorten group names for x-axis and display
        group_agg['Group Short'] = group_agg['Group'].apply(lambda x: x.split('/')[-1] if '/' in x else x)
        group_agg['Avg Score'] = group_agg['Avg Score'].astype(int)
        # Reorder columns for display
        cols = ['Group Short', 'User Count', 'Avg Score', 'Min Score', 'Max Score', 'Group']
        group_agg = group_agg[cols]
        group_agg = group_agg.sort_values('Avg Score', ascending=False)
        print(f"{Fore.GREEN}✓ Created aggregations for {len(group_agg)} groups{Style.RESET_ALL}")
        return group_agg

class ExcelReportGenerator:
    @staticmethod
    def create_report(df, group_df, output_filename):
        print("\n📝 Creating Excel report...")
        wb = Workbook()

        # Sheet 1: User Activity Report
        ws_users = wb.active
        ws_users.title = "User Activity Report"
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws_users.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws_users.cell(row=row_idx, column=col_idx, value=value)
        ExcelReportGenerator._adjust_column_widths(ws_users)

        # Sheet 2: Group Summary
        group_display_cols = ['Group Short', 'User Count', 'Avg Score', 'Min Score', 'Max Score']
        ws_groups = wb.create_sheet("Group Summary")
        for col_idx, header in enumerate(group_display_cols, 1):
            cell = ws_groups.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_idx, row in enumerate(group_df[group_display_cols].itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws_groups.cell(row=row_idx, column=col_idx, value=value)
        ExcelReportGenerator._adjust_column_widths(ws_groups)

        # Clean Bar Chart: Average Experience Score by Group
        if group_df.shape[0] > 0:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Group-wise Average Experience Score"
            chart.y_axis.title = "Average Experience Score"
            chart.x_axis.title = "Group"
            chart.style = 2
            data = Reference(ws_groups, min_col=3, min_row=1, max_col=3, max_row=group_df.shape[0] + 1)
            cats = Reference(ws_groups, min_col=1, min_row=2, max_row=group_df.shape[0] + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.legend = None
            chart.width = 16
            chart.height = 8
            chart.gapWidth = 100
            ws_groups.add_chart(chart, "H2")

        wb.save(output_filename)
        print(f"✓ Report saved to: {output_filename}")

    @staticmethod
    def _adjust_column_widths(worksheet):
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

def select_time_range():
    print("\nSelect report time range:")
    print("1. Last 24 hours")
    print("2. Last 7 days")
    print("3. Last 30 days")
    print("4. Custom (max 30 days, format YYYY-MM-DD HH:MM)")
    while True:
        choice = input("Enter your choice [1-4]: ").strip()
        if choice in {"1", "2", "3", "4"}:
            break
        print("Invalid input. Please select 1, 2, 3 or 4.")
    now = datetime.now()
    if choice == "1":
        return now - timedelta(days=1), now
    elif choice == "2":
        return now - timedelta(days=7), now
    elif choice == "3":
        return now - timedelta(days=30), now
    elif choice == "4":
        fmt = "%Y-%m-%d %H:%M"
        while True:
            s = input("Start datetime (YYYY-MM-DD HH:MM): ")
            e = input("End datetime   (YYYY-MM-DD HH:MM): ")
            try:
                sdt = datetime.strptime(s, fmt)
                edt = datetime.strptime(e, fmt)
                if edt <= sdt:
                    print("End time must be after start time.")
                    continue
                if (edt - sdt).days > 30:
                    print("Range cannot be more than 30 days.")
                    continue
                return sdt, edt
            except Exception as ex:
                print("Invalid date/time format. Try again.")

def select_timezone():
    print("\nSelect time zone for input/output:")
    print("a. UTC/GMT (recommended for API)")
    print("b. Local machine time")
    while True:
        choice = input("Enter your choice [a/b]: ").strip().lower()
        if choice in {"a", "b"}:
            break
        print("Invalid input. Please select a or b.")
    return "utc" if choice == "a" else "local"

def main():
    print(f"\n{Fore.MAGENTA}{'='*60}{Style.RESET_ALL}")
    print(f"{Fore.MAGENTA}{'Netskope User Report Generator':^60}{Style.RESET_ALL}")
    print(f"{Fore.MAGENTA}{'='*60}{Style.RESET_ALL}")

    tz_choice = select_timezone()
    sdt, edt = select_time_range()
    if tz_choice == "utc":
        sdt = sdt.replace(tzinfo=timezone.utc)
        edt = edt.replace(tzinfo=timezone.utc)
    else:
        sdt = sdt.astimezone(timezone.utc)
        edt = edt.astimezone(timezone.utc)

    parsed_url = urlparse(API_URL)
    instance_name = parsed_url.netloc
    print(f"\n{Fore.CYAN}⚙️  Configuration:{Style.RESET_ALL}")
    print(f"   • Instance: {instance_name}")
    print(f"   • API URL: {API_URL}")
    print(f"   • Time Range: {sdt} to {edt} (UTC)")
    print(f"   • Sort Order: {SORT_ORDER} (by expScore)")
    print(f"   • Max records per API call: 100 (48 hr window)")
    if API_TOKEN == "YOUR_API_TOKEN_HERE":
        print(f"\n{Fore.RED}❌ Error: Please update the API_TOKEN at the top of the script{Style.RESET_ALL}")
        return

    client = NetskopeAPIClient(API_URL, API_TOKEN)
    try:
        total_start_time = time.time()
        users = client.get_all_users_chunked(sdt, edt, sort_order=SORT_ORDER)
        if not users:
            print(f"\n{Fore.YELLOW}⚠️  No users found for the specified time range{Style.RESET_ALL}")
            return
        processor = UserDataProcessor()
        df = processor.process_users(users)
        if df.empty:
            print(f"\n{Fore.YELLOW}⚠️  No users with Experience Score > 0 in this period.{Style.RESET_ALL}")
            return
        group_df = processor.create_group_aggregation(df)
        print(f"\n{Fore.CYAN}📊 Data Summary:{Style.RESET_ALL}")
        print(f"   • Total unique users: {len(df):,}")
        print(f"   • Unique email addresses: {df['User Email'].nunique():,}")
        print(f"   • Score range: {df['Experience Score'].min():.1f} - {df['Experience Score'].max():.1f}")
        print(f"   • Average score: {df['Experience Score'].mean():.1f}")
        print(f"   • Unique groups: {len(group_df):,}")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_filename = f"netskope_users_{timestamp}.xlsx"
        report_generator = ExcelReportGenerator()
        report_generator.create_report(df, group_df, output_filename)
        total_elapsed = time.time() - total_start_time
        print(f"\n{Fore.GREEN}{'='*60}{Style.RESET_ALL}")
        print(f"{Fore.GREEN}✓ Report generation completed successfully!{Style.RESET_ALL}")
        print(f"{Fore.GREEN}📊 Total execution time: {total_elapsed:.1f} seconds{Style.RESET_ALL}")
        print(f"{Fore.GREEN}📄 Output file: {output_filename}{Style.RESET_ALL}")
        print(f"{Fore.GREEN}{'='*60}{Style.RESET_ALL}\n")
    except KeyboardInterrupt:
        print(f"\n\n{Fore.YELLOW}⚠️  Process interrupted by user{Style.RESET_ALL}")
    except Exception as e:
        print(f"\n{Fore.RED}❌ Error generating report: {e}{Style.RESET_ALL}")
        raise

if __name__ == "__main__":
    main()
