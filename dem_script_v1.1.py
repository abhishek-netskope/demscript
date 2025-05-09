import requests
import time
import logging
from datetime import datetime, timedelta, timezone
from collections import defaultdict
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

API_URL = "https://eliterhythm.goskope.com/api/v2/adem/users/getentities"
API_TOKEN = "13342187c927b00bdfa826e31218f"
LIMIT = 100
MAX_WINDOW_SECONDS = 48 * 3600
TIME_RANGES = {'1day': 1, '7days': 7, '30days': 30}

logging.basicConfig(filename='netskope_run.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

errors, api_durations = [], []

def datetime_to_epoch(dt):
    return int(dt.timestamp())

def clean_group_name(name):
    if name and '/' in name:
        return name.split('/')[-1]
    return name

def fetch_data(starttime, endtime, total_calls, current_call):
    offset, all_users = 0, []
    while True:
        payload = {"starttime": starttime, "endtime": endtime, "limit": LIMIT, "offset": offset}
        call_start = time.time()
        try:
            response = requests.post(API_URL, headers={"Authorization": f"Bearer {API_TOKEN}",
                                                       "Content-Type": "application/json"},
                                     json=payload, verify=False)  # 🔧 PATCHED LINE
        except Exception as e:
            errors.append(str(e))
            logging.error(str(e))
            break
        duration = time.time() - call_start
        api_durations.append(duration)

        if response.status_code == 429:
            retry_after = int(response.headers.get('Retry-After', '1'))
            errors.append(f"429 Too Many Requests. Waiting {retry_after}s...")
            time.sleep(retry_after)
            continue
        if 500 <= response.status_code < 600:
            errors.append(f"Server error {response.status_code}")
            time.sleep(5)
            continue
        if response.status_code != 200:
            errors.append(f"Error {response.status_code}: {response.text}")
            break

        data = response.json()
        users = data.get('users', [])
        if not users:
            break
        all_users.extend(users)

        offset += LIMIT
        total = data.get('totalUsersCount', 0)
        progress_pct = (current_call[0] / total_calls) * 100
        avg_dur = sum(api_durations) / len(api_durations)
        eta = avg_dur * (total_calls - current_call[0])
        print(f"Progress: {progress_pct:.1f}% — ETA: {timedelta(seconds=int(eta))}")
        logging.info(f"Page {offset}, duration={duration:.2f}s, progress={progress_pct:.1f}%")

        current_call[0] += 1
        if offset >= total:
            break
        time.sleep(1)
    return all_users

def aggregate_users(raw_users):
    agg = defaultdict(lambda: {'userGroups': set(), 'expScores': [], 'applications': set(),
                               'devices': set(), 'locations': set()})
    for user in raw_users:
        u = user.get('user', '')
        agg[u]['userGroups'].update(filter(None, user.get('userGroups', [])))
        exp = user.get('expScore')
        if exp is not None:
            agg[u]['expScores'].append(exp)
        agg[u]['applications'].update(filter(None, user.get('applications', [])))
        agg[u]['devices'].update([d.get('deviceName', '') for d in user.get('devices', []) if d.get('deviceName')])
        loc = user.get('location', '')
        if loc:
            agg[u]['locations'].add(loc)

    final_data = []
    for user, details in agg.items():
        avg_expScore = round(sum(details['expScores']) / len(details['expScores'])) if details['expScores'] else ''
        apps = sorted(details['applications'])
        devs = sorted(details['devices'])
        locs = sorted(details['locations'])
        groups = sorted(details['userGroups']) if details['userGroups'] else ['']
        for group in groups:
            final_data.append({'user': user,
                               'userGroup': clean_group_name(group),
                               'expScore': avg_expScore,
                               'applications': ','.join(apps),
                               'applicationsCount': len(apps),
                               'devices': ','.join(devs),
                               'locations': ','.join(locs)})
    return pd.DataFrame(final_data)

def save_to_excel(df_users, filename_xlsx):
    df_users['expScore'] = pd.to_numeric(df_users['expScore'], errors='coerce').round(0).astype('Int64')
    df_avg = df_users.groupby("userGroup").agg(
        average_expScore=("expScore", "mean"),
        user_count=("user", "count")
    ).reset_index()
    df_avg['average_expScore'] = df_avg['average_expScore'].round(0).astype('Int64')
    df_chart = df_avg.sort_values(by="average_expScore", ascending=False)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "User Data"
    ws2 = wb.create_sheet("Group_Averages")
    ws3 = wb.create_sheet("Group_Chart")

    for sheet, df in zip([ws1, ws2, ws3], [df_users, df_avg, df_chart]):
        for r in dataframe_to_rows(df, index=False, header=True):
            sheet.append(r)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")
            cell.alignment = Alignment(horizontal="center")
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = max_len + 2
            for cell in col:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))

    chart = BarChart()
    chart.title = "Average Experience Score by Group (Sorted)"
    chart.y_axis.title = "User Groups"
    chart.x_axis.title = "Experience Score"
    chart.type = "bar"
    chart.style = 10
    chart.shape = 4
    data_ref = Reference(ws3, min_col=2, min_row=1, max_row=ws3.max_row)
    cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=ws3.max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws3.add_chart(chart, "E5")

    wb.save(filename_xlsx)

def fetch_and_save(days, filename_xlsx):
    now, end, start = datetime.now(tz=timezone.utc), datetime.now(tz=timezone.utc), datetime.now(tz=timezone.utc) - timedelta(days=days)
    all_raw_data, total_chunks = [], (days * 86400) // MAX_WINDOW_SECONDS + 1
    total_calls, current_call = total_chunks * 200, [1]

    while end > start:
        chunk_start = max(end - timedelta(seconds=MAX_WINDOW_SECONDS), start)
        s_epoch, e_epoch = datetime_to_epoch(chunk_start), datetime_to_epoch(end)
        print(f"📦 Fetching {chunk_start} → {end}")
        logging.info(f"Chunk {chunk_start} → {end}")
        all_raw_data.extend(fetch_data(s_epoch, e_epoch, total_calls, current_call))
        end = chunk_start

    df_users = aggregate_users(all_raw_data)
    save_to_excel(df_users, filename_xlsx)

    print("\n=== Run Summary ===")
    print(f"Total API calls: {len(api_durations)}")
    print(f"Avg API call time: {sum(api_durations)/len(api_durations):.2f}s" if api_durations else "No calls")
    print(f"Total errors: {len(errors)}")
    for err in errors:
        print(f"⚠️ {err}")

def main():
    print("Select time range to pull:\n1 → last 1 day\n2 → last 7 days\n3 → last 30 days")
    choice = input("Enter choice (1/2/3): ").strip()
    if choice == '1':
        label, days = '1day', 1
    elif choice == '2':
        label, days = '7days', 7
    elif choice == '3':
        label, days = '30days', 30
    else:
        print("❌ Invalid choice.")
        return
    fetch_and_save(days, f'users_last_{label}_cleaned.xlsx')

if __name__ == '__main__':
    main()
